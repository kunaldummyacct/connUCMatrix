from flask import Flask, render_template, request, redirect, session, jsonify
import pandas as pd
import os
from functools import wraps

app = Flask(__name__)
app.secret_key = os.urandom(24)

# Look for Excel file in the same directory as this script
EXCEL_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'ConnUCMatrix.xlsx')
ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', 'admin123')  # Set via environment variable or use default

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect('/login')
        return f(*args, **kwargs)
    return decorated

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        login_type = request.form.get('login_type')
        
        if login_type == 'guest':
            session['logged_in'] = True
            session['user_type'] = 'guest'
            return redirect('/')
        elif login_type == 'admin':
            if request.form.get('password') == ADMIN_PASSWORD:
                session['logged_in'] = True
                session['user_type'] = 'admin'
                return redirect('/')
            return render_template('login.html', error='Invalid admin password')
        
        return render_template('login.html', error='Please select login type')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.pop('logged_in', None)
    return redirect('/login')

@app.route('/')
@login_required
def index():
    user_type = session.get('user_type', 'guest')
    return render_template('index.html', user_type=user_type)

@app.route('/api/data')
@login_required
def get_data():
    from openpyxl import load_workbook
    
    # Check if Excel file exists
    if not os.path.exists(EXCEL_FILE):
        return jsonify({
            'error': f'Excel file not found: {EXCEL_FILE}',
            'message': 'Please place ConnUCMatrix.xlsx in the same directory as app.py'
        }), 404
    
    # Load workbook to get merge info and actual cell values
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    # Get merged cell ranges
    merged_cells = []
    for merged_range in ws.merged_cells.ranges:
        merged_cells.append({
            'min_row': merged_range.min_row,
            'max_row': merged_range.max_row,
            'min_col': merged_range.min_col,
            'max_col': merged_range.max_col
        })
    
    # Read headers directly from openpyxl to preserve values
    headers = []
    for row_idx in range(1, 4):  # Rows 1, 2, 3
        row_data = []
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            row_data.append(cell_value if cell_value is not None else '')
        headers.append(row_data)
    
    # Read data starting from row 4
    data_df = pd.read_excel(EXCEL_FILE, header=2)
    data_df = data_df.fillna('')
    columns = list(data_df.columns)
    
    return jsonify({
        'headers': headers,
        'columns': columns,
        'data': data_df.values.tolist(),
        'merged_cells': merged_cells
    })

@app.route('/api/update', methods=['POST'])
@login_required
def update_data():
    # Check if user is admin
    if session.get('user_type') != 'admin':
        return jsonify({'error': 'Unauthorized: Admin access required'}), 403
    
    from openpyxl import load_workbook
    
    payload = request.json
    columns = payload['columns']
    data = payload['data']
    
    # Load existing workbook to preserve headers and merged cells
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    # Save merged cell ranges from header rows (rows 1-3)
    header_merges = [merge for merge in ws.merged_cells.ranges if merge.min_row <= 3]
    
    # Delete all existing data rows (keep only 3 header rows)
    if ws.max_row > 3:
        ws.delete_rows(4, ws.max_row - 3)
    
    # Re-apply header merged cells (in case they were affected)
    for merge in header_merges:
        if merge not in ws.merged_cells:
            ws.merge_cells(str(merge))
    
    # Insert new data rows (starting from row 4, after 3 header rows)
    for row_idx, row_data in enumerate(data, start=4):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    wb.save(EXCEL_FILE)
    return jsonify({'success': True})

if __name__ == '__main__':
    app.run(debug=True, port=5000)
